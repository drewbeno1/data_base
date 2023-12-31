import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl
# import sv_ttk
import pandas as pd

# Define a variable to store the selected file path
selected_file_path = None
# Define treeview as a global variable
treeview = None

# Create a list to store the data and corresponding Excel row indices
data_and_indices = []

# Create a dictionary to store Pitch_Count for each Pitcher
pitch_count_dict = {}

def load_data():
    global selected_file_path, treeview  # Declare global variables
    # Show a file dialog to select the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not file_path:
        return  # User canceled the file dialog
    
    for item in treeview.get_children():
        treeview.delete(item)

    selected_file_path = file_path  # Store the selected file path
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Sheet1']
    list_values = list(sheet.values)
    for col_name in list_values[0]:
        # Adjust the column name to match your Excel sheet
        treeview.heading(col_name, text=col_name)
    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)

    # Check if the file is empty or contains data
    if len(list_values) > 0:
        # If it contains data, find the last row and its Pitch_Count
        last_row = list_values[-1]
        last_pitcher = last_row[0]
        last_pitch_count = last_row[2]  # Assuming Pitch_Count is in the third column (index 2)
        
        # Store the last Pitch_Count for the last Pitcher
        pitch_count_dict[last_pitcher] = last_pitch_count


def insert_row():
    global selected_file_path, treeview, pitch_count_dict, last_pitcher  # Declare global variables
    global last_pitcher  # Declare last_pitcher as a global variable
    if selected_file_path is None:
        # Prompt the user to load a file first
        tk.messagebox.showerror("Error", "Please load a file first.")
        return

    Pitcher = Pitcher_Combobox.get()
    Date = Date_entry.get()
    Velo = velo_entry.get()
    Type = type_Combobox.get()
    Result = Result_Combobox.get()

    # Validate the ComboBoxes and Entry fields
    if (
        Pitcher not in name_list
        or Date == "Date"
        or Velo == "Velo"
        or Type not in pitch_type_list
        or Result not in pitch_result_list
    ):
        tk.messagebox.showerror("Get it right bro", "Please fill in all fields correctly")
        return
    
    # Check if it's a new pitcher or a repeat
    if Pitcher != last_pitcher:
        # Initialize the Pitch_Count for the current Pitcher if it's a new Pitcher
        pitch_count_dict[Pitcher] = 0

    # Increment the Pitch_Count for the current Pitcher
    pitch_count_dict[Pitcher] += 1
    Pitch_Count = pitch_count_dict[Pitcher]
    
    # Insert row into Excel sheet using the selected file path
    workbook = openpyxl.load_workbook(selected_file_path)
    sheet = workbook['Sheet1']
    row_values = [Pitcher, Date, Pitch_Count, Velo, Type, Result]
    sheet.append(row_values)
    workbook.save(selected_file_path)
    # Insert row into treeview
    treeview.insert('', tk.END, values=row_values)
    treeview.see(treeview.get_children()[-1])
    # Clear the values
    velo_entry.delete(0, "end")
    velo_entry.insert(0, "Velo")
    Result_Combobox.delete(0, "end")
    Result_Combobox.insert(0, "Result")
    type_Combobox.delete(0, "end")
    type_Combobox.insert(0, "Pitch Type")

    # Update the last_pitcher variable # This stores the last pitcher off so that it knows to restart counting or not
    last_pitcher = Pitcher

# Initialize the last_pitcher variable to None
last_pitcher = None

def remove_row():
    global treeview, selected_file_path
    selected_items = treeview.selection()
    if not selected_items:
        return
    if selected_file_path is None:
        tk.messagebox.showerror("Error", "Please load a file first.")
        return
    # Load the workbook
    workbook = openpyxl.load_workbook(selected_file_path)
    sheet = workbook['Sheet1']
    # Iterate over selected items and delete corresponding rows in the Excel file
    for item in selected_items:
        item_values = treeview.item(item, 'values')
        # Find the row index of the item in the Excel file based on the Date column (adjust as needed)
        date_to_delete = item_values[1]  # Assuming Date is the second column (index 1)
        pitcher = item_values[0]
        if pitcher in pitch_count_dict:
            pitch_count_dict[pitcher] -= 1  # Decrement the pitch count for the removed row
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
            if row[0].value == date_to_delete:
                sheet.delete_rows(row[0].row)
    # Save the updated workbook
    workbook.save(selected_file_path)
    # Remove selected items from the treeview
    for item in selected_items:
        treeview.delete(item)



def enter_key_pressed(event):
    insert_row()


root = tk.Tk()
root.title('Bloomsburg Baseball Pitching Charts')

# YOU CANT STYLE BUTTON BACKGROUNDS WITH SV_TTK :(
style = ttk.Style(root)
# sv_ttk.set_theme("light") 
style.configure('elder.TButton')
style.map('elder.TButton', background=[('active', '#007fff')])

name_list = ['Andrew Armstrong', 'Nate Baranski', 'Kolby Barrow', 'Mike Cacioppo', 'Jack Carver', 
             'Cole Coolbaugh', 'Dominic Coombe', 'Nick Heubel', 'Dansby Koppisch', 'Jake Kuperavage',
             'Tyler LePage', 'Landon Lorson', 'Dylan Lubinski', 'Josh Marquard', 'Emmet McLaughlin',
             'Travis Peden', 'Kaden Peifer', 'Zach Steen', 'Xander Velez', 'Matt Vernieri', 'Brian Walsh',
             'Owen Wilhide', 'Christian Zito', 'Will Dean', 'Scott Gilbert', 'Mason Keene', 'James Scott', 'Mike Standen']
pitch_result_list = ["Ball", "Strike looking", "Strike swinging", "Foul Ball", "Strikeout looking", "Strikeout swinging",
                     "BIP Out", "Single", "Double", "Triple", "HR", "Walk", "HBP", "Drop 3rd & Safe"]
pitch_type_list = ['FB', 'CB', 'SL', 'CH', 'Splitter', 'Cutter', 'Knuck', 'Eephus']

frame = ttk.Frame(root)
frame.grid(row=0, column=0, padx=20, pady=10, sticky="nsew")
frame.grid_rowconfigure(0, weight=1)
frame.grid_rowconfigure(1, weight=0) 
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Insert Data")
widgets_frame.grid(row=0, column=0, padx=20, pady=30)

Pitcher_Combobox = ttk.Combobox(widgets_frame, values=name_list)
Pitcher_Combobox.current(0)  # Set the initial value by index
Pitcher_Combobox.set("Pitcher")  # Set the placeholder text
Pitcher_Combobox.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

Date_entry = ttk.Entry(widgets_frame)
Date_entry.insert(0, "mm/dd/yyyy")
Date_entry.bind("<FocusIn>", lambda e: Date_entry.delete('0', 'end'))
Date_entry.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

# We are gonna find pitch count manually
# count_spinbox = ttk.Spinbox(widgets_frame, from_=0, to=200)
# count_spinbox.insert(0, "Pitch Counter")
# count_spinbox.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

velo_entry = ttk.Entry(widgets_frame)
velo_entry.insert(0, "Velo")
velo_entry.bind("<FocusIn>", lambda e: velo_entry.delete('0', 'end'))
velo_entry.grid(row=4, column=0, padx=5, pady=(0, 5), sticky="ew")

type_Combobox = ttk.Combobox(widgets_frame, values=pitch_type_list)
type_Combobox.current(0)  # Set the initial value by index
type_Combobox.set("Pitch Type")  # Set the placeholder text
type_Combobox.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

Result_Combobox = ttk.Combobox(widgets_frame, values=pitch_result_list)
Result_Combobox.current(0)  # Set the initial value by index
Result_Combobox.set("Pitch Result")  # Set the placeholder text
Result_Combobox.grid(row=6, column=0, padx=5, pady=5, sticky="ew")

button = ttk.Button(widgets_frame, text="Enter Pitch", command=insert_row, style='elder.TButton')
button.grid(row=7, column=0, padx=5, pady=5, sticky="nsew")
button.bind('<Return>', enter_key_pressed)
button.bind('<KP_Enter>', enter_key_pressed) 

separator = ttk.Separator(widgets_frame)
separator.grid(row=2, column=0, padx=(20, 10), pady=10, sticky="ew")

# Button to remove a row
remove_button = ttk.Button(widgets_frame, text="Remove Selected Row", command=remove_row)
remove_button.grid(row=8, column=0, padx=5, pady=(0, 5), sticky="ew")
remove_button.bind('<Return>', enter_key_pressed)
remove_button.bind('<KP_Enter>', enter_key_pressed) 

# frame for excel data
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")



# Define your column identifiers
cols = ("Pitcher", "Date", "Pitch Count", "Velo", "Pitch Type", "Result")

treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=cols, height=13)
treeview.column("Pitcher", width=100)
treeview.column("Date", width=100)
treeview.column("Pitch Count", width=100)
treeview.column("Velo", width=100)
treeview.column("Pitch Type", width=100)
treeview.column("Result", width=100)

treeview.pack()
treeScroll.config(command=treeview.yview)
load_data()

root.mainloop()


#  ####### App closes and data cleaning happens for analysis ##########

# Load and edit the existing Excel workbook
# ##################### Create a by pitch sheet #####################
workbook = openpyxl.load_workbook(selected_file_path)
sheet = workbook['Sheet1']
df = pd.DataFrame(sheet.values)
df.columns = df.iloc[0]
df = df[1:]
# df.index = None
# print(df)

# Let's go ahead and create a primary key just in case we need it 
# df['Pitch Key'] =  df['Pitcher'] + '-' + df['Date'] + '-' + df['Pitch Counter']

# Find extra columns
df['Strike or Ball'] = df['Result'].apply(lambda x: 'Ball' if x in ['Ball', 'Walk', 'HBP'] else 'Strike')

result_to_swing = {
    'Ball': 'No swing',
    'Walk': 'No swing',
    'HBP': 'No swing',
    'Strike looking': 'No swing',
    'Strikeout looking': 'No swing',
    'Foul Ball': 'Swing contact',  # Note: Removed the extra double quotes around 'Foul Ball'
    'Single': 'Swing contact',
    'Double': 'Swing contact',
    'Triple': 'Swing contact',
    'HR': 'Swing contact',
    'BIP Out': 'Swing contact',
    'Strike swinging': 'Swing no contact',
    'Drop 3rd & Safe': 'Swing no contact',
    'Strikeout swinging': 'Swing no contact'
}
df['Swing'] = df['Result'].map(result_to_swing)

result_of_ab = {
    'Ball': 'nothing',
    'Walk': 'free base',
    'HBP': 'free base',
    'Strike looking': 'nothing',
    'Strikeout looking': 'not free base',
    'Foul Ball': 'nothing',  # Note: Removed the extra double quotes around 'Foul Ball'
    'Single': 'not free base',
    'Double': 'not free base',
    'Triple': 'not free base',
    'HR': 'not free base',
    'BIP Out': 'not free base',
    'Strike swinging': 'nothing',
    'Drop 3rd & Safe': 'not free base',
    'Strikeout swinging': 'not free base'
}
df['Free Bases'] = df['Result'].map(result_of_ab)

Event_or_no = {
    'Walk': 'event',
    'HBP': 'event',
    'Strikeout looking': 'event',
    'Single': 'event',
    'Double': 'event',
    'Triple': 'event',
    'HR': 'event',
    'BIP Out': 'event',
    'Drop 3rd & Safe': 'event',
    'Strikeout swinging': 'event'
}
df['Event'] = df['Result'].map(Event_or_no).fillna('not event')


# ######## Create Count #########
# Temporarily make the df smaller so that we can work on it like an actual df. then we'll add that filler column back in

mydf2 = df
mydf2 = mydf2.reset_index()

# ## Step 1: Specify when a new ab occurs. 
# Create a new column to store the modified index
mydf2['new_index'] = mydf2.index
# Initialize a counter for 'new ab' occurrences. Start at 1 so that first row can be 1
ab_counter = 1

# Iterate through the DataFrame rows. 
# ### The new atbats column will tell us when to loop through and find the counts and restart
for i, row in mydf2.iterrows():
    if row['Event'] == 'event':
        next_row_index = i + 1
        if next_row_index < len(mydf2): 
            ab_counter += 1
            mydf2.at[next_row_index, 'ab'] = f'ab {ab_counter}'

mydf2.at[0, 'ab'] = 'ab 1'
# Fill down 
mydf2['ab'].fillna(method='ffill', inplace=True)
# Drop extra index columns 
mydf2 = mydf2.drop(columns={'index', 'new_index'})
# print(mydf2)

# ### Now that we have at bat trackers, for each at bat count the pitches delayed. 
# We will create a "Strike or Ball 2" column that we will seperate out as its own df
# Then create a column called index in the original df and
# Then add a null to the top row and move every other field down by 1 index in the new s or b df
# Then reset that index and join it back in. 
# Now every pitch can be counted at 1 pitch later
# Then just need logic to make the strike or ball field null each row right after an event and for it to restart 
#    the count each ab 
mydf3 = mydf2

mydf4 = mydf3[['Strike or Ball']]
# The new null row to add
new_row = pd.Series({'Strike or Ball': 'dummy'})
# Insert the new row at the specified position
mydf4 = pd.concat([new_row, mydf4], ignore_index=True)
# Add in index row & rename Strike or Ball
mydf4 = mydf4.reset_index()
mydf4 = mydf4.rename(columns={'Strike or Ball': 'Strike or Ball Count Tracker'})

# Now add in index for original df 
mydf3 = mydf3.reset_index()

# Now merge new one back in. It worked woo! 
mydf3 = mydf3.merge(mydf4[['index', 'Strike or Ball Count Tracker']], on='index', how='left')

mydf3 = mydf3.drop(columns='index')

# ## Now we can do what we were trying to do before with the pandas counter

# Set the first pitch of each ab to not be tracked since that is just telling us the pitch outcome of the ab anyway we dont care
# Find the first occurrence of each value in 'ab'
first_occurrence_mask = ~mydf3['ab'].duplicated()

# Set the 'Strike or Ball Count Tracker' value to null for the first occurrences
mydf3.loc[first_occurrence_mask, 'Strike or Ball Count Tracker'] = None

# Initialize 'Balls' and 'Strikes' columns with 0
# Initialize variables to keep track of the cumulative count of balls and strikes
# Initialize 'Balls' and 'Strikes' columns with 0
mydf3['Balls'] = 0
mydf3['Strikes'] = 0

# Track counts based on 'ab' and 'Strike or Ball'
current_ab = None
ball_count = 0
strike_count = 0

for index, row in mydf3.iterrows():
    if current_ab != row['ab']:
        # Start a new 'ab' group
        current_ab = row['ab']
        ball_count = 0
        strike_count = 0
    
    if row['Strike or Ball Count Tracker'] == 'Ball':
        ball_count += 1
    elif row['Strike or Ball Count Tracker'] == 'Strike':
        strike_count += 1
    
    mydf3.at[index, 'Balls'] = ball_count
    mydf3.at[index, 'Strikes'] = strike_count

# mydf3.to_excel('test.xlsx', index=False)

# Now create 'Count' Row by Balls - Strikes # THIS ISNT WORKING
mydf3['Count'] = mydf3['Balls'].astype(str) + '-' + mydf3['Strikes'].astype(str)



df = mydf3
df = df[0:]
# print(df)

# print(df.info())
# Assign this to analysis sheet and include extra columns
analysis_sheet = workbook['pitch breakdown']
# Clear all rows except the header row (assuming header is in the first row)
analysis_sheet.delete_rows(analysis_sheet.min_row + 1, analysis_sheet.max_row)
# Write the manipulated data from the DataFrame to the analysis sheet
for index, row in df.iterrows():
    analysis_sheet.append(row.tolist())


# Save the updated workbook
workbook.save(selected_file_path)

# ################# Create a by player sheet with their stats #####################
workbook = openpyxl.load_workbook(selected_file_path)
sheet = workbook['pitch breakdown']
df2 = pd.DataFrame(sheet.values)
df2.columns = df2.iloc[0]
df2[1:]

# print(df2)

# avg FB Velo
fb_df = df2[['Pitcher', 'Velo', 'Pitch Type']]
fb_df = fb_df[fb_df['Pitch Type'] == 'FB']
fb_df['Velo'] = fb_df['Velo'].astype('Int64')
avg_fb = fb_df.groupby('Pitcher')['Velo'].mean().reset_index().round(1)
avg_fb = avg_fb.rename(columns={'Velo': 'avg FB'})
avg_fb = avg_fb.fillna(0)

# Top FB Velo
top_fb = fb_df.groupby('Pitcher')['Velo'].max().reset_index().round(1)
top_fb = top_fb.rename(columns={'Velo': 'Top FB'})
top_fb = top_fb.fillna(0)

# Strike % 
# Filter 'df' to only include rows where 'Pitch result' is 'Strike'
strike_df = df2[df2['Strike or Ball'] == 'Strike']
# Group the filtered DataFrame by 'Pitcher' and calculate strike percentage
strike_percentage = (strike_df.groupby('Pitcher')['Strike or Ball'].count() / df2.groupby('Pitcher')['Strike or Ball'].count()).round(2)
strike_percentage = strike_percentage.fillna(0)
strike_percentage = strike_percentage.reset_index().rename(columns={'Strike or Ball': 'Strike %'})
strike_percentage = strike_percentage[strike_percentage['Pitcher'] != 'Pitcher']
# print(strike_percentage)

# Whiff %
# Filter dfs for Swings 
swing_df = df2[df2['Swing'] != "No swing"]
whiff_df = df2[df2['Swing'] == 'Swing no contact']
# Group by Pitcher
whiff_percentage = (whiff_df.groupby('Pitcher')['Swing'].count() / swing_df.groupby('Pitcher')['Swing'].count()).round(2)
whiff_percentage = whiff_percentage.fillna(0)
whiff_percentage = whiff_percentage.reset_index().rename(columns={'Swing': 'Whiff %'})
whiff_percentage = whiff_percentage[whiff_percentage['Pitcher'] != 'Pitcher']


# Total CSW %
called_or_whiff_count = df2[df2['Result'].isin(['Strike looking', 
                                                'Strike swing & miss', 
                                                'Drop 3rd & Safe', 
                                                'Strikeout looking', 
                                                'Strikeout swinging'])].groupby('Pitcher')['Result'].count()
total_results_count = df2.groupby('Pitcher')['Result'].count()
CSW = ((called_or_whiff_count) / total_results_count).round(2)
CSW = CSW.fillna(0)
CSW = CSW.reset_index().rename(columns={'Result': 'CSW % '})
CSW = CSW[CSW['Pitcher'] != 'Pitcher']

# FB CS+W % & Offspeed CSW %  ((Called Strikes + Swings and misses)/number of pitches)
# ## FB 
fb_csw = df2[df2['Pitch Type'] == 'FB']
called_or_whiff_count = fb_csw[fb_csw['Result'].isin(['Strike looking', 
                                                      'Strike swing & miss', 
                                                      'Drop 3rd & Safe', 
                                                      'Strikeout looking', 
                                                      'Strikeout swinging'])].groupby('Pitcher')['Result'].count()
total_results_count = fb_csw.groupby('Pitcher')['Result'].count()
fb_CSW = (called_or_whiff_count / total_results_count).round(2)
fb_CSW = fb_CSW.fillna(0)
fb_CSW = fb_CSW.reset_index().rename(columns={'Result': 'FB CSW %'})
fb_CSW = fb_CSW[fb_CSW['Pitcher'] != 'Pitcher']

# ## OFF SPEED
os_csw = df2[df2['Pitch Type'] != 'FB']
called_or_whiff_count = os_csw[os_csw['Result'].isin(['Strike looking', 
                                                      'Strike swing & miss', 
                                                      'Drop 3rd & Safe', 
                                                      'Strikeout looking', 
                                                      'Strikeout swinging'])].groupby('Pitcher')['Result'].count()
total_results_count = os_csw.groupby('Pitcher')['Result'].count()
os_CSW = (called_or_whiff_count / total_results_count).round(2)
os_CSW = os_CSW.fillna(0)
os_CSW = os_CSW.reset_index().rename(columns={'Result': 'OffSpeed CSW %'})
# os_CSW = os_CSW[fb_CSW['Pitcher'] != 'Pitcher']
# print(os_CSW)

# Free base count
free_df = df2[df2['Free Bases'] == 'free base']
free_bases = (free_df.groupby('Pitcher')['Free Bases'].count()).round(1)

# Reformat df2 to be just unique Pitchers and then join their data back in 
df2 = df2[['Pitcher']]
df2 = df2.drop_duplicates()
df2 = df2.merge(avg_fb, how='left', on='Pitcher')
df2 = df2.merge(top_fb, how='left', on='Pitcher')
df2 = df2.merge(strike_percentage, how='left', on='Pitcher')
df2 = df2.merge(whiff_percentage, how='left', on='Pitcher')
df2 = df2.merge(CSW, how='left', on='Pitcher')
df2 = df2.merge(fb_CSW, how='left', on='Pitcher')
df2 = df2.merge(os_CSW, how='left', on='Pitcher')
df2 = df2.merge(free_bases, how='left', on='Pitcher')
df2 = df2.fillna(0)

# print(df2)

# Setup new page on excel file
# Assign this to analysis sheet and include extra columns
pitcher_sheet = workbook['pitcher breakdown']

# Clear and load it back into Sheet1
pitcher_sheet.delete_rows(pitcher_sheet.min_row + 1, pitcher_sheet.max_row)
# Write the manipulated data from the DataFrame to the analysis sheet
for index, row in df2.iterrows():
    pitcher_sheet.append(row.tolist())


workbook.save(selected_file_path)

# Close the Excel file
workbook.close()