import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl
import sv_ttk
import pandas as pd

# Define a variable to store the selected file path
selected_file_path = None
# Define treeview as a global variable
treeview = None
# Create a list to store the data and corresponding Excel row indices
data_and_indices = []

def load_data():
    global selected_file_path, treeview  # Declare global variables
    # Show a file dialog to select the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not file_path:
        return  # User canceled the file dialog
    
    for item in treeview.get_children():
        treeview.delete(item)

    selected_file_path = file_path  # Store the selected file path
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Sheet1']
    list_values = list(sheet.values)
    for col_name in list_values[0]:
        # Adjust the column name to match your Excel sheet
        treeview.heading(col_name, text=col_name)
    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)


def insert_row():
    global selected_file_path, treeview  # Declare global variables
    if selected_file_path is None:
        # Prompt the user to load a file first
        tk.messagebox.showerror("Error", "Please load a file first.")
        return

    Pitcher = Pitcher_Combobox.get()
    Date = Date_entry.get()
    Pitch_Count = count_spinbox.get()
    Velo = velo_entry.get()
    Type = type_Combobox.get()
    Result = Result_Combobox.get()
    # Insert row into Excel sheet using the selected file path
    workbook = openpyxl.load_workbook(selected_file_path)
    sheet = workbook['Sheet1']
    row_values = [Pitcher, Date, Pitch_Count, Velo, Type, Result]
    sheet.append(row_values)
    workbook.save(selected_file_path)
    # Insert row into treeview
    treeview.insert('', tk.END, values=row_values)
    # Clear the values
    velo_entry.delete(0, "end")
    velo_entry.insert(0, "Velo")
    Result_Combobox.delete(0, "end")
    Result_Combobox.insert(0, "Result")
    type_Combobox.delete(0, "end")
    type_Combobox.insert(0, "Pitch Type")

def remove_row():
    global treeview, selected_file_path
    selected_items = treeview.selection()
    if not selected_items:
        return
    if selected_file_path is None:
        tk.messagebox.showerror("Error", "Please load a file first.")
        return
    # Load the workbook
    workbook = openpyxl.load_workbook(selected_file_path)
    sheet = workbook['Sheet1']
    # Iterate over selected items and delete corresponding rows in the Excel file
    for item in selected_items:
        item_values = treeview.item(item, 'values')
        # Find the row index of the item in the Excel file based on the Date column (adjust as needed)
        date_to_delete = item_values[1]  # Assuming Date is the second column (index 1)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
            if row[0].value == date_to_delete:
                sheet.delete_rows(row[0].row)
    # Save the updated workbook
    workbook.save(selected_file_path)
    # Remove selected items from the treeview
    for item in selected_items:
        treeview.delete(item)



def enter_key_pressed(event):
    insert_row()


root = tk.Tk()
root.title('Bloomsberg Baseball Pitching Charts')

style = ttk.Style(root)
sv_ttk.set_theme("light") 
style.configure("TLabel", foreground="red")    # Set foreground (text) color to red
style.configure("TButton", foreground="black")   # Set button text color to blue
style.configure("Custom.TSpinbox",
                arrowsize=20,  # Adjust the arrow size as needed
                arrowcolor="black")
# sv_ttk.set_theme("dark")

root.tk_setPalette(background='#ececec')

name_list = ["Sammy", "Woody", "Argo", "Epstein"]
pitch_result_list = ["Strike looking", "Strike swing & miss", "Foul Ball", "Ball", "Strikeout looking", "Strikeout swinging",
                     "BIP Out", "Hit", "Walk", "HBP", "Drop 3rd & Safe"]
pitch_type_list = ['FB', 'CB', 'SL', 'CH', 'Splitter', 'Cutter', 'Knuck', 'Eephus']

frame = ttk.Frame(root)
frame.grid(row=0, column=0, padx=20, pady=10, sticky="nsew")
frame.grid_rowconfigure(0, weight=1)
frame.grid_rowconfigure(1, weight=0) 
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Insert Data")
widgets_frame.grid(row=0, column=0, padx=20, pady=30)

Pitcher_Combobox = ttk.Combobox(widgets_frame, values=name_list)
Pitcher_Combobox.current(0)  # Set the initial value by index
Pitcher_Combobox.set("Pitcher")  # Set the placeholder text
Pitcher_Combobox.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

Date_entry = ttk.Entry(widgets_frame)
Date_entry.insert(0, "Date")
Date_entry.bind("<FocusIn>", lambda e: Date_entry.delete('0', 'end'))
Date_entry.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

count_spinbox = ttk.Spinbox(widgets_frame, from_=0, to=200)
count_spinbox.insert(0, "Pitch Counter")
count_spinbox.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

velo_entry = ttk.Entry(widgets_frame)
velo_entry.insert(0, "Velo")
velo_entry.bind("<FocusIn>", lambda e: velo_entry.delete('0', 'end'))
velo_entry.grid(row=4, column=0, padx=5, pady=(0, 5), sticky="ew")

type_Combobox = ttk.Combobox(widgets_frame, values=pitch_type_list)
type_Combobox.current(0)  # Set the initial value by index
type_Combobox.set("Pitch Type")  # Set the placeholder text
type_Combobox.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

Result_Combobox = ttk.Combobox(widgets_frame, values=pitch_result_list)
Result_Combobox.current(0)  # Set the initial value by index
Result_Combobox.set("Pitch Result")  # Set the placeholder text
Result_Combobox.grid(row=6, column=0, padx=5, pady=5, sticky="ew")

button = ttk.Button(widgets_frame, text="Enter", command=insert_row)
button.grid(row=7, column=0, padx=5, pady=5, sticky="nsew")
button.bind('<Return>', enter_key_pressed)
button.bind('<KP_Enter>', enter_key_pressed) 

separator = ttk.Separator(widgets_frame)
separator.grid(row=2, column=0, padx=(20, 10), pady=10, sticky="ew")

# Button to remove a row
remove_button = ttk.Button(widgets_frame, text="Remove Selected Row", command=remove_row)
remove_button.grid(row=8, column=0, padx=5, pady=(0, 5), sticky="ew")
remove_button.bind('<Return>', enter_key_pressed)
remove_button.bind('<KP_Enter>', enter_key_pressed) 

# frame for excel data
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")



# Define your column identifiers
cols = ("Pitcher", "Date", "Pitch Counter", "Velo", "Pitch Type", "Result")

treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=cols, height=13)
treeview.column("Pitcher", width=100)
treeview.column("Date", width=100)
treeview.column("Pitch Counter", width=100)
treeview.column("Velo", width=100)
treeview.column("Pitch Type", width=100)
treeview.column("Result", width=100)

treeview.pack()
treeScroll.config(command=treeview.yview)
load_data()

root.mainloop()



#  ####### App closes and data cleaning happens for analysis ##########

# Load and edit the existing Excel workbook
# ##################### Create a by pitch sheet #####################
workbook = openpyxl.load_workbook(selected_file_path)
sheet = workbook['Sheet1']
df = pd.DataFrame(sheet.values)
df.columns = df.iloc[0]

# Find extra columns
df['Strike or Ball'] = df['Result'].apply(lambda x: 'Ball' if x in ['Ball', 'Walk', 'HBP'] else 'Strike')
result_to_swing = {
    'Ball': 'No swing',
    'Walk': 'No swing',
    'HBP': 'No swing',
    'Strike looking': 'No swing',
    'Strikeout looking': 'No swing',
    'Foul Ball': 'Swing contact',  # Note: Removed the extra double quotes around 'Foul Ball'
    'Hit': 'Swing contact',
    'BIP Out': 'Swing contact',
    'Strike swing & miss': 'Swing no contact',
    'Drop 3rd & Safe': 'Swing no contact',
    'Strikeout swinging': 'Swing no contact'
}
df['Swing'] = df['Result'].map(result_to_swing)

# Assign this to analysis sheet and include extra columns
analysis_sheet = workbook['pitch breakdown']

# Clear and load it back into Sheet1
analysis_sheet.delete_rows(analysis_sheet.min_row, analysis_sheet.max_row)
# Write the manipulated data from the DataFrame to the analysis sheet
for index, row in df.iterrows():
    analysis_sheet.append(row.tolist())

####include extra column names in the right index
analysis_sheet.cell(row=1, column=7, value='Strike or Ball')
analysis_sheet.cell(row=1, column=8, value='Swing')

# Save the updated workbook
workbook.save(selected_file_path)

# ################# Create a by player sheet #####################
workbook = openpyxl.load_workbook(selected_file_path)
sheet = workbook['pitch breakdown']
df2 = pd.DataFrame(sheet.values)



# Close the Excel file
workbook.close()